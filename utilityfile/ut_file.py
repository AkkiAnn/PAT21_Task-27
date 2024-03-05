import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill

class EXL:

    def count_row(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_row)

    def count_column(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_coulmn)

    def read_data(file, sheetName, rownum, columnno):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(rownum, columnno).value

    def write_data(file, sheetName, rownum, columnno, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(rownum, columnno).value = data
        workbook.save(file)

    def fill_green_color(file, sheetName, rownum, columnno):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        greenFill = PatternFill(start_color='60b212',
                                end_color='60b212',
                                fill_type='solid')
        sheet.cell(rownum, columnno).fill = greenFill
        workbook.save(file)

    def fill_red_color(file, sheetName, rownum, columnno):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        redFill = PatternFill(start_color='ff0000',
                              end_color='ff0000',
                              fill_type='solid')
        sheet.cell(rownum, columnno).fill = redFill
        workbook.save(file)